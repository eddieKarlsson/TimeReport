import openpyxl as xl
import sys


def insert_data_to_report(wb_path, days_dict, usr_s_dict):
    # Try to open excel
    try:
        wb = xl.load_workbook(wb_path)
    except FileNotFoundError as e:
        print(f'Error! {wb_path} file not found, will exit', e)
        sys.exit()

    #  Set sheet
    ws = wb.active

    # Insert data to sheet
    start_day_column = 12
    for i, day in enumerate(days_dict, start=start_day_column):
        if day["hours"] is not None:
            h = day["hours"]
        else:
            h = 0

        if day["overtime_1"] is not None:
            h += day["overtime_1"]

        if day["overtime_2"] is not None:
            h += day["overtime_2"]

        ws.cell(column=i, row=10).value = h
        ws.cell(column=i, row=11).value = day["travel_time"]

    ws['M5'] = usr_s_dict['year']
    week = usr_s_dict['week']
    #  ws['L12'] = round(week / 4.348)
    ws['L5'] = week

    wb.save(wb_path)
