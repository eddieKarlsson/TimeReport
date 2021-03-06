import openpyxl as xl
import sys


def insert_data_to_report(wb_path, days_dict, usr_s_dict, week_sheetname='W'):
    # Try to open excel
    try:
        wb = xl.load_workbook(wb_path)
    except FileNotFoundError as e:
        print(f'Error! {wb_path} file not found, will exit', e)
        sys.exit()

    #  Set sheet
    ws = wb.active

    # Insert data to sheet
    start_day_row = 15
    for i, day in enumerate(days_dict, start=start_day_row):
        ws.cell(row=i, column=4).value = day["hours"]
        ws.cell(row=i, column=5).value = day["overtime_1"]
        ws.cell(row=i, column=6).value = day["overtime_2"]
        ws.cell(row=i, column=7).value = day["travel_time"]

    ws['J12'] = usr_s_dict['projno']  # Projnr
    ws['L12'] = usr_s_dict['year']  # Year
    week = usr_s_dict['week']

    month = round(week / 4.348)
    if month < 1:
        ws['M12'] = 1
    else:
        ws['M12'] = month

    ws['N12'] = week

    ws.title = week_sheetname  # set sheet name

    wb.save(wb_path)
