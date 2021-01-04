import shutil
import os
import openpyxl as xl
import sys

from settings import Settings


class CustomerReport():
    """Create a customer timreport from internal reporting excels"""

    def __init__(self):
        # Create settings object and load settings from JSON
        self.s = Settings()
        self.usr_s = self.s.load_user_settings()

    def open_in_excel(self):
        # Concetenate file path to excel month report
        in_wb_path = os.path.join(self.usr_s['in_time_folder'],
                                  self.usr_s['in_time_file'])
        # Open Excel file
        try:
            self.in_wb = xl.load_workbook(in_wb_path, data_only=True)
        except FileNotFoundError as e:
            print(f'Error! {self.in_wb_path} file not found, will exit', e)
            sys.exit()

        # Open Sheet
        try:
            self.sheet = self.in_wb[self.usr_s['month']]
        except KeyError:
            print(f'Error! {self.sheet} sheet does not exist, will exit')
            sys.exit()

    def _in_excel_find_week(self):
        # Loop A column, until week is found
        for i in range(7, 300):
            cell = self.sheet.cell(row=i, column=1)
            if str(self.usr_s['week']) in str(cell.value):
                self.start_row = i
                start_row_found = True
                break

        if not start_row_found:
            print(f'Start row not found in sheet {self.sheet}, will exit')
            sys.exit()

    def _in_excel_find_days_create_dict(self):
        """Loop in B column to find the days in the week"""
        self.dayCounter = 0
        self.days = []  # Create empty dict

        for i in range(self.start_row, 150):
            cell_a = self.sheet.cell(row=i, column=1)
            cell_b = self.sheet.cell(row=i, column=2)

            if self.dayCounter > 6:
                break

            # If we have counted 2 days and there is a value in
            # A column, guess its a new week and break
            if self.dayCounter >= 2 and cell_a.value is not None:
                break

            if cell_b.value is None:
                continue
            else:  # day found, print day and add corresponding data to dict
                self.dayCounter += 1
                dateStr = str(cell_b.value)

                # Create dictionary
                day = {
                    "day_index": None,
                    "date": dateStr,
                    "start_row": i,
                    "stop_row": None,
                    "hours": None,
                    "travel_time": None,
                    "overtime_1": None,
                    "overtime_2": None,
                }

                self.days.append(day)

    def _in_excel_find_day_data(self):
        projnr = self.usr_s['projno']

        """ Set variable stop_row by accesing next days start_row - 1 """
        for index, day in enumerate(self.days):
            day["day_index"] = index

            if index < 6:  # if it's the last day of the week dont get stop_row
                try:
                    nextDay = self.days[index + 1]
                    nextDayStartRow = nextDay["start_row"]
                    day["stop_row"] = nextDayStartRow - 1
                except IndexError:
                    print(f'WARNING: Seems like week only have {index} days')
            elif index == 6:
                day["stop_row"] = day["start_row"]

        """ check if project-nr is in any of the rows """
        for day in self.days:
            #  print(day)
            if day["stop_row"] is not None:
                for i in range(day["start_row"], day["stop_row"]):

                    cellP = self.sheet.cell(row=i, column=4)
                    cellPVal = str(cellP.value)
                    if cellP.value is not None and str(projnr) in cellPVal:
                        cell = self.sheet.cell(row=i, column=7)
                        if cell.value is not None:
                            day["travel_time"] = cell.value

                        cell = self.sheet.cell(row=i, column=8)
                        if cell.value is not None:
                            day["hours"] = cell.value

                        cell = self.sheet.cell(row=i, column=9)
                        if cell.value is not None:
                            day["overtime_1"] = cell.value

                        cell = self.sheet.cell(row=i, column=10)
                        if cell.value is not None:
                            day["overtime_2"] = cell.value

            if day["start_row"] == day["stop_row"]:
                i = day["start_row"]
                cellP = self.sheet.cell(row=i, column=4)
                if cellP.value is not None and str(projnr) in str(cellP.value):
                    cell = self.sheet.cell(row=i, column=7)
                    day["travel_time"] = cell.value
                    cell = self.sheet.cell(row=i, column=8)
                    day["hours"] = cell.value
                    cell = self.sheet.cell(row=i, column=9)
                    day["overtime_1"] = cell.value
                    cell = self.sheet.cell(row=i, column=10)
                    day["overtime_2"] = cell.value

    def in_excel_print_created_dic(self):
        for day in self.days:
            print("\t", day)
        print('\n')

    def user_input_choose_template(self):
        print("SELECT TIME REPORT:")

        time_template = {
            1: {'Filename': 'MC SV.xlsx', 'type': 'mc', 'lang': 'sv'},
            2: {'Filename': 'MC ENG.xlsx', 'type': 'mc', 'lang': 'en'},
            3: {'Filename': 'TP SPAIN.xlsx', 'type': 'tp_spain', 'lang': 'en'}
            }
        for key, value in time_template.items():
            print("\t", key, ' : ', value)

        select = int(input("Choose timereport: "))

        self.ttp = time_template[select]

        print(self.ttp)

    def out_excel_copy_template(self):
        """Copy template to output folder and rename"""
        template_file = 'data/template/' + self.ttp['Filename']
        t_file, t_ext = os.path.splitext(template_file)

        # Create new file name, change some strings if Swedish
        if self.ttp['lang'] == 'sv':
            t = 'Tidrapport - '
            self.w = 'V'
        else:
            t = 'Timereport - '
            self.w = 'W'
        self.w += str(self.usr_s['week'])  # add week number to string
        self.new_report = (self.usr_s['output_folder'] + t
                           + str(self.usr_s['projno']) + ' - ' + self.w
                           + t_ext)

        shutil.copy(template_file, self.new_report)
        print('\t', self.new_report)

    def out_excel_write_data(self):
        """Select which template to import and then execute"""

        self.week_format = self.w

        if self.ttp['type'] == 'mc':
            from mc import insert_data_to_report
            insert_data_to_report(self.new_report, self.days, self.usr_s,
                                  week_sheetname=self.week_format)

        if self.ttp['type'] == 'tp_spain':
            from tp_spain import insert_data_to_report
            insert_data_to_report(self.new_report, self.days, self.usr_s)

    def run(self):
        self.open_in_excel()
        self._in_excel_find_week()
        self._in_excel_find_days_create_dict()
        self._in_excel_find_day_data()
        self.in_excel_print_created_dic()
        self.user_input_choose_template()
        self.out_excel_copy_template()
        self.out_excel_write_data()


if __name__ == '__main__':
    app = CustomerReport()
    app.run()
